from openpyxl import load_workbook

class xlsxInput :
    path = None
    sheet = None
    data = None

    def __init__(self, path)    :
        self.path = path
        self.sheet = self.takeInputFromExcel()
        self.data = self.prossingData()

    def takeInputFromExcel(self) :   
        workbook = load_workbook(self.path, data_only=True)
        sheet = workbook.active
        return sheet

    def prossingData(self) :
        arrayData = []
        for row in self.sheet.iter_rows(values_only=True):
            tempArray = []
            for cell in row:
                tempArray.append(cell if cell is not None else '')
            arrayData.append(tempArray)
        return arrayData

    def display(self) :
        for i in range(0, len(self.data))   :
            print(self.data[i])

